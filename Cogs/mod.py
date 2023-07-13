import openpyxl
import datetime
from discord import File
wb = openpyxl.load_workbook('/home/gihyeon/bot/challenge.xlsx')
load_ws = wb['Sheet1']


def Today_challenge():
    d = datetime.datetime.now()
    str_now = str(d.strftime("%y-%m-%d\n"))
    if d.month % 2 == 0:
        col = 2
        send_message = (str_now + (load_ws.cell(d.day, col).value))
        wb.close()
    else:
        col = 1
        send_message = (str_now + (load_ws.cell(d.day, col).value))
        wb.close()
    return send_message


def Today_file():
    d = datetime.datetime.now()
    if d.month % 2 == 0:
        col = 2
        file = File(f'/home/gihyeon/bot/{col}/{d.day}.jpg', filename="image.jpg")
        wb.close()
    else:
        col = 1
        file = File(f"/home/gihyeon/bot/{col}/{d.day}.jpg", filename="image.jpg")
        wb.close()
    return file


def Tomorrow_challenge():
    d = datetime.datetime.now()
    tomorrow = d + datetime.timedelta(days=1)
    strftime_tomorrow = int(tomorrow.strftime("%d"))
    str_tomorrow = str(tomorrow.strftime("%y-%m-%d\n"))
    if tomorrow.month % 2 == 0:
        col = 2
        send_message = (
            str_tomorrow + (load_ws.cell(strftime_tomorrow, col).value))
    else:
        col = 1
        send_message = (
            str_tomorrow + (load_ws.cell(strftime_tomorrow, col).value))
    return send_message


def Tomorrow_file():
    d = datetime.datetime.now()
    tomorrow = d + datetime.timedelta(days=1)
    strftime_tomorrow = int(tomorrow.strftime("%d"))
    if tomorrow.month % 2 == 0:
        col = 2
        file = File(
            f"./{col}/{strftime_tomorrow}.jpg", filename="image.jpg")
        wb.close()
    else:
        col = 1
        file = File(
            f"./{col}/{strftime_tomorrow}.jpg", filename="image.jpg")
        wb.close()
    return file
